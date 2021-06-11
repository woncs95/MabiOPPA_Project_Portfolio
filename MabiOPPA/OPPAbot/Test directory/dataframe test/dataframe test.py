import pandas as pd
from pprint import pprint
import asyncio
from openpyxl import load_workbook

dic={'dics':{'x':{'a':2},'y':{'b':3},'z':{'c':3}}}
dic['sth']='everything'
dic['sth']='anything'
#print(dic)

#a=['a','b','c']
#print(a[2])

async def excel(dic):
    xl=load_workbook(filename='dic.xlsx', data_only=True)
    ws=xl['Sheet1']
    for cell in ws['1']:  # 1행의 모든 셀을 확인
        print(cell.value)
    df = pd.DataFrame(dic)
    pprint(df)
    df.to_excel('dic.xlsx')


async def main():
    await excel(dic)

asyncio.run(main())
